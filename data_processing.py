"""
Paid items
IBK:318+560
Developed by: Adam Gruchacz, PAT
Process Expert: Piotr Kolanowski
13/04/2023
"""
from datetime import datetime
from typing import Optional
import re
import shutil

from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import openpyxl
import pandas as pd

import messagebox

'#Global variables '
FINAL_INFORMATION: str = ""
SH_INTER_DOUBLE: str = 'InterDouble'
SH_CROSS_DOUBLE: str = 'CrossDouble'
PATH_OUTPUT_DIRECTORY: str = ''
IDYES = 6


def get_sheetnames_xlsx(filepath: str, sheetindex: int) -> str:
    """
     function gets sheet name basing on index
     :param filepath: string of path.
     :param sheetindex: integer representing tab index, which starts from 0
     :return: str
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_links=False)
    sheetname = wb.sheetnames[sheetindex]
    return sheetname


def rename_sheet(filepath: str, sheetname: str, expectedsheetname: str) -> str:
    """
    function open wb, renames sheet`s tab, and save a file
    :param filepath: string of path.
    :param sheetname: string representing existing sheet name
    :param expectedsheetname: string representing needed tab name
    :return: str
    """

    global FINAL_INFORMATION
    try:
        wb = openpyxl.load_workbook(filepath)
        wb_sheetname = wb[sheetname]
        wb_sheetname.title = expectedsheetname
        wb.save(filepath)
    except Exception:
        FINAL_INFORMATION = "error occured in function 'rename sheet '"
    return FINAL_INFORMATION


def remove_unneccesary_characters(originalinvoice: str) -> str:
    """
    function passes list for removing characters from string
    :param originalinvoice: string represenging invoice including redundant characters
    :return: str
    """

    invoice = str(originalinvoice).upper()
    ''''#2 backslashes means backslash because one backslash means escaping'''
    characters_replacement(["-", "/", ".", ",", " ", "'", "", "`", "´", "_", "—", "#", "\\", '"'], 0, invoice)
    characters_replacement(["IDM", "ICM", "OOP"], 0, invoice)
    characters_replacement(["AS", "SA", "AAA", "SSS", "AA", "SS", "A", "S"], 1, invoice)
    return invoice


def characters_replacement(characterslist: list, ending: int, invoice: str) -> str:
    """
        function removes unneccesary characters from string
        :param characterslist: list represenging phrases to be removed
        :param ending: int - 0 for replacement, 1 for replacement if phrase on end
        :param invoice: string represenging invoice including redundant characters
        :return: str
    """
    for character in characterslist:
        if ending == 0:
            invoice = invoice.replace(character, "")
        elif ending == 1:
            invoice = re.sub(character + "$", '', invoice)
    return invoice


def combine_dataframes(tuple_paths: tuple, selected_erp: str, df_paid: pd.DataFrame) -> pd.DataFrame:
    """
    function combines data from paid files in one dataframe
    :param tuple_paths: list including all - selected in Gui by user - paid files
    :param selected_erp: str variable including Dynamics or PeopleSoft value
    :param df_paid: dataframe for paid invoices
    :return: DataFrame including data from paid file(s)
    """

    global FINAL_INFORMATION
    if selected_erp == 'Dynamics':
        'Dynamics columns'
        listcolumns = ["Account No", "Invoice", "Vendor Name", "Gross Amount"]
    elif selected_erp == 'PeopleSoft':
        listcolumns = ["Account", "Invoice Number", "Supplier", "Gross Invoice Amount"]
    try:
        # lst of column names which needs to be string
        lst_str_cols = ['Invoice', 'Invoice Number']
        # use dictionary comprehension to make dict of dtypes
        dict_dtypes = {x: 'str' for x in lst_str_cols}
        counter = 0
        for path in tuple_paths:
            counter += 1
            if counter == 1:
                if selected_erp == 'Dynamics':
                    '''# lst of column names which needs to be string
                    lst_str_cols = ['Invoice']
                    # use dictionary comprehension to make dict of dtypes
                    dict_dtypes = {x: 'str' for x in lst_str_cols}'''
                    # use dict on dtypes
                    df_paid = pd.read_excel(path, dtype=dict_dtypes)
                elif selected_erp == 'PeopleSoft':
                    df_paid = pd.read_excel(path, skiprows=1)
                    '''df_paid.drop('Supplier', inplace=True, axis=1)'''
                    # renaming due to 2 columns named 'Supplier' in input file
                    '''df_paid = df_paid.rename(columns={'Supplier.1': 'Supplier'})'''
                df_paid = pd.DataFrame(data=df_paid, columns=listcolumns)
            else:
                if selected_erp == 'Dynamics':
                    df_temp = pd.read_excel(path, dtype=dict_dtypes)
                elif selected_erp == 'PeopleSoft':
                    df_temp = pd.read_excel(path, skiprows=1)
                    df_temp.drop('Supplier', inplace=True, axis=1)
                    df_temp = df_temp.rename(columns={'Supplier.1': 'Supplier'})
                df_temp = pd.DataFrame(data=df_temp, columns=listcolumns)
                df_paid = pd.concat([df_paid, df_temp], axis=0)  # merge dataframes from files
        if selected_erp == 'Dynamics': # removing leading zeros from Supplier & invoice fields
            df_paid['Invoice'] = df_paid['Invoice'].str.lstrip('0')# 2022-2-7
            '''df_paid['Account No'] = df_paid['Account No'].str.lstrip('0')'''
        elif selected_erp == 'PeopleSoft':
            df_paid['Invoice Number'] = df_paid['Invoice Number'].str.lstrip('0')  # 2022-2-7
            '''df_paid['Supplier'] = df_paid['Supplier'].str.lstrip('0')'''
        df_paid.columns = ['Account', 'Invoice Number', 'Supplier', 'Gross Invoice Amount']
    except Exception:
        FINAL_INFORMATION = "dataframes not merged. " + 'Error in combine_dataframes'
    return df_paid


def get_dataframe_to_pay(pathtobepaid: str, erp: str) -> pd.DataFrame:
    """
   function creates dataframe basing on file to be paid
   :param pathtobepaid: path of to be paid file
   :param erp: string equals 'Dynamics' or 'PeopleSoft
   :return: DataFrame including data from 'to pay' file
    """
    global FINAL_INFORMATION
    'Dynamics columns'
    if erp == 'Dynamics':
        listcolumns = ["Vendor Account", "Invoice", "Name", "Payment amount"]
    elif erp == 'PeopleSoft':
        listcolumns = ["Remit Supplier", "Invoice", "Supplier Name", "Paid Amount"]
    try:
        # lst of column names which needs to be string
        lst_str_cols = ['Invoice', 'Invoice Number']
        # use dictionary comprehension to make dict of dtypes
        dict_dtypes = {x: 'str' for x in lst_str_cols}
        df_topay = pd.read_excel(pathtobepaid, dtype=dict_dtypes)
        bSupplierAccount = False
        for col in df_topay.columns:
            if col =='Supplier account':
                listcolumns = ["Supplier account", "Invoice", "Name", "Payment amount"]
                bSupplierAccount = True
                break
        df_topay = pd.DataFrame(data=df_topay, columns=listcolumns)
        if erp == 'PeopleSoft':
            df_topay = df_topay.rename(columns={'Remit Supplier': 'Supplier account', 'Supplier Name': 'Name',
                                                'Paid Amount': 'Payment amount'})
            return df_topay
        elif bSupplierAccount == False:
            df_topay = df_topay.rename(columns={'Vendor Account': 'Supplier account'})
            return df_topay
        '''else:
            # alrernative column name of Supplier Account
            listcolumns = ["Supplier account", "Invoice", "Name", "Payment amount"]
            df_topay = pd.DataFrame(data=df_topay, columns=listcolumns)'''
    except Exception:
        FINAL_INFORMATION = "not read data for to pay. " + "error in get_dataframe_to_pay"
    return df_topay


def modify_dataframe(df: pd.DataFrame, accountcolumnname: str, erp: str) -> Optional[pd.DataFrame]:
    """
     function add to dataframe columns InvoiceConverted and INV_ACCOUNT
     :param df: respective dataframe (can be representing paid or to be paid)
     :param accountcolumnname: string representing Account/Supplier column
     :param erp: string variable  including value PeopleSoft or Dynamics
     :return: if no error function return DataFrame; else return None
    """
    global FINAL_INFORMATION
    try:
        if accountcolumnname == 'Supplier' and erp == "PeopleSoft":
            column_invoice = 'Invoice Number'
        if accountcolumnname == 'Account':
            column_invoice = 'Invoice Number'
        elif erp == "Dynamics" or accountcolumnname == 'Supplier account':
            column_invoice = 'Invoice'
        else:
            column_invoice = 'Invoice Number'
        df = df[df[column_invoice].notna()]
        #7/2/2022  ->removing leading zeros in invoice number
        df[column_invoice] = df[column_invoice].str.lstrip('0')
        df['InvoiceConverted'] = df.apply(lambda row: remove_unneccesary_characters(row[column_invoice]), axis=1)
        if erp == "Dynamics":
            df["INV_ACCOUNT"] = df['InvoiceConverted'].astype(str) + "_" + df[accountcolumnname].astype(str)
            '''df["INV_ACCOUNT"] = df['InvoiceConverted'].astype(str) + "_" + df[accountcolumnname]'''
        elif erp == "PeopleSoft":
            if accountcolumnname == 'Supplier account':
                df[accountcolumnname] = df[accountcolumnname].astype(str).str.split('.').str[0]
            else:
                df[accountcolumnname] = df[accountcolumnname].str.lstrip('0')
            df["INV_ACCOUNT"] = df['InvoiceConverted'].astype(str) + "_" + df[accountcolumnname].astype(str)
            # df["TEST"] = df['Invoice'].astype(str) + "A"
    except Exception:
        FINAL_INFORMATION = "An error occurred in function: modify_dataframe. "
        return None
    return df


def update_cell_and_change_column_width(ws: openpyxl.worksheet.worksheet.Worksheet, columnname: str, width: int,
                                       headername: str, rownumber: str) -> None:
    """
    function put text to respective cell and adjust width of column
    :param ws: respective worksheet
    :param columnname: string representing of column letter(s)
    :param width: integer representing expected width of column
    :param headername: string representing needed column header
    :param rownumber: string representing number of row for headers
    :return: None
     """
    ws[columnname + rownumber].value = headername
    ws.column_dimensions[columnname].width = width


def prepare_inter_double(wb: openpyxl.workbook.workbook.Workbook, df_topay: pd.DataFrame) -> str:
    """
   function makes needed adjustment in Dataframe and export dataframe to Excel to InterDouble tab
   :param wb: workbook
   :param df_topay: dataframe including data from file "to pay"
   :return: string variable if error happened; else empty string
    """

    global FINAL_INFORMATION
    global SH_INTER_DOUBLE
    try:
        # keep=False means show all rows including duplicates
        df_duplicates = df_topay.loc[df_topay["INV_ACCOUNT"].duplicated(keep=False), :]
        df_duplicates.drop(['InvoiceConverted'], axis=1)
        df_duplicates.index += 2  # ' update index to have row number'
        df_duplicates = df_duplicates[['Supplier account', 'Name', 'Invoice', 'Payment amount', 'INV_ACCOUNT']]
        df_duplicates = df_duplicates.sort_values(by=['INV_ACCOUNT'])
        wb.create_sheet(SH_INTER_DOUBLE)
        ws = wb[SH_INTER_DOUBLE]
        rows = dataframe_to_rows(df_duplicates, index=True, header=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx+1, column=c_idx, value=value)
        ws["A1"].value = "To be Paid"
        update_cell_and_change_column_width(ws, 'A', 8, "Rows Id", "2")
        update_cell_and_change_column_width(ws, 'B', 15, 'Supp #', "2")
        update_cell_and_change_column_width(ws, 'C', 45, 'Supp Name', "2")
        update_cell_and_change_column_width(ws, 'D', 15, 'Inv Num', "2")
        update_cell_and_change_column_width(ws, 'E', 12, 'Amm', "2")
        update_cell_and_change_column_width(ws, 'F', 20, 'Invoice+Supplier Id', "2")

    except Exception:
        FINAL_INFORMATION = "An error occurred in function: prepare_Inter_Double. "
    return FINAL_INFORMATION


def prepare_cross_double(wb: openpyxl.workbook.workbook.Workbook, df_cross_double: pd.DataFrame, filepath: str, erp: str) -> str:
    """
   function makes needed adjustment in Dataframe and export dataframe to Excel to CrossDouble tab
   :param wb: workbook
   :param df_cross_double: dataframe including joined data from paid files and to be paid file
   :param filepath: string representing path of output file
   :return: string variable if error happened; else empty string
    """

    global FINAL_INFORMATION
    global SH_CROSS_DOUBLE
    try:
        if erp == "Dynamics":
            account_header = 'Account'
        elif erp == "PeopleSoft":
            account_header = 'Supplier'
        # remove in next 2 lines when Account No is empty
        df_cross_double[account_header].replace('', np.nan, inplace=True)
        df_cross_double.dropna(subset=[account_header], inplace=True)
        wb.create_sheet(SH_CROSS_DOUBLE)
        ws = wb[SH_CROSS_DOUBLE]
        # order of columns
        df_cross_double = df_cross_double[['Supplier account', 'Name', 'Invoice', 'INV_ACCOUNT',
                                           'Payment amount', account_header, 'Supplier', 'Invoice Number',
                                           'Gross Invoice Amount']]
        df_cross_double = df_cross_double.sort_values(by=['Supplier account'])
        df_cross_double = df_cross_double.drop_duplicates()
        rows = dataframe_to_rows(df_cross_double, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        ws.insert_cols(6)   # empty column between to be paid  and paid
        update_cell_and_change_column_width(ws, 'A', 15, 'Supp #(To pay)', "1")
        update_cell_and_change_column_width(ws, 'B', 30, 'Supp Name', "1")
        update_cell_and_change_column_width(ws, 'C', 10, 'Inv Num', "1")
        update_cell_and_change_column_width(ws, 'D', 20, 'Invoice+Supplier Id', "1")
        update_cell_and_change_column_width(ws, 'E', 12, 'Amm', "1")
        update_cell_and_change_column_width(ws, 'F', 6, '', "1")
        update_cell_and_change_column_width(ws, 'G', 15, 'Supp #(Paid)', "1")
        update_cell_and_change_column_width(ws, 'H', 30, 'Supp Name', "1")
        update_cell_and_change_column_width(ws, 'I', 10, 'Inv Num', "1")
        update_cell_and_change_column_width(ws, 'J', 12, 'Amm', "1")
        wb.save(filepath)

    except Exception:
        FINAL_INFORMATION = "An error occurred in function: prepare_Cross_Double. "
    return FINAL_INFORMATION


def process(path_to_be_paid: str, paid_tuple: tuple, outputfolder: str, selected_erp: str) -> None:
    """
    function gets paths from GUI, keep logic of tool flow and display final output message
    :param path_to_be_paid: string representing path of 'to be paid' file
    :param paid_tuple: list including path(s) of 'paid' file(s)
    :param outputfolder: string representing directory path where output file will be stored
    :param selected_erp: string representing selected erp system
    :return: None
     """
    global FINAL_INFORMATION
    df_paid = pd.DataFrame()
    df_paid = combine_dataframes(paid_tuple, selected_erp, df_paid)
    if selected_erp == "Dynamics":
        df_paid = modify_dataframe(df_paid, 'Account', selected_erp)
    elif selected_erp == "PeopleSoft":
        df_paid = modify_dataframe(df_paid, 'Supplier', selected_erp)
    if df_paid is not None:
        df_topay = get_dataframe_to_pay(path_to_be_paid, selected_erp)
        df_topay = modify_dataframe(df_topay, 'Supplier account', selected_erp)
        if df_topay is not None:
            output_dir_path = f"{outputfolder}\DoublePayments {datetime.now():%d%m%Y %H%M%S}.xlsx"
            shutil.copy(path_to_be_paid, output_dir_path)
            firstsheetname = get_sheetnames_xlsx(output_dir_path, 0)
            if rename_sheet(output_dir_path, firstsheetname, "To be Paid") == '':
                wb = openpyxl.load_workbook(output_dir_path)
                if prepare_inter_double(wb, df_topay) == '':
                    # remove if double invoices+Account in the df to pay
                    df_topay = df_topay.drop_duplicates(subset='INV_ACCOUNT', keep="last")
                    df_cross_double = pd.merge(df_topay, df_paid, how="inner", on=["INV_ACCOUNT"])
                    prepare_cross_double(wb, df_cross_double, output_dir_path, selected_erp)
    if not FINAL_INFORMATION:
        FINAL_INFORMATION = "Completed"

    messagebox.showinformation('Final result', FINAL_INFORMATION)
    exit()
